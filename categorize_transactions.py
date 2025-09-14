import pandas as pd
import sys
from openpyxl import load_workbook

# --- Script Configuration ---
# Set the filenames for your two Excel files here.
# Make sure they are in the same directory as this script, or provide the full path.
MASTER_FILE = 'master_transactions.xlsx'
CATEGORIES_FILE = 'Categories.xlsx'

# Set the name of the sheet in the Categories file that contains the mapping data.
CATEGORIES_SHEET_NAME = 'Transactions'

# Define the columns to be used for matching and populating.
# These names must exactly match the column headers in your Excel files.
MATCH_COLUMN = 'Transaction Description'
CATEGORY_COLUMN = 'Category'
SUBCATEGORY_COLUMN = 'Subcategory'

# --- Main Script Logic ---
def categorize_transactions_with_openpyxl(master_file, categories_file, categories_sheet_name,
                                         match_col, cat_col, subcat_col):
    """
    Reads two Excel files, finds the first matching transaction description,
    and updates the category and subcategory columns in the master file
    without changing the file's original formatting.
    """
    try:
        # Step 1: Read the categories mapping file using pandas for a fast lookup.
        print(f"Reading categories file: '{categories_file}' (sheet: '{categories_sheet_name}')...")
        df_categories = pd.read_excel(categories_file, sheet_name=categories_sheet_name)

        # Step 2: Create a dictionary for a fast lookup.
        # We drop duplicates and keep the first occurrence to ensure we only
        # use the first match, as requested.
        print("Creating a lookup dictionary from category data...")
        category_map = df_categories.drop_duplicates(
            subset=[match_col], keep='first'
        ).set_index(match_col).to_dict('index')

        # Step 3: Load the master workbook using openpyxl.
        print(f"Loading master file with openpyxl: '{master_file}'...")
        workbook = load_workbook(master_file)
        sheet = workbook.active

        # Step 4: Get column headers to find correct indices.
        header_row = [cell.value for cell in sheet[1]]
        try:
            match_col_idx = header_row.index(match_col) + 1
            cat_col_idx = header_row.index(cat_col) + 1
            subcat_col_idx = header_row.index(subcat_col) + 1
        except ValueError as e:
            print(f"Error: Column '{e}' not found in the master file header.", file=sys.stderr)
            sys.exit(1)
            
        print("Updating Category and Subcategory columns...")
        
        # Step 5: Iterate through the rows and update the cells.
        for row_idx in range(2, sheet.max_row + 1):
            transaction_desc = sheet.cell(row=row_idx, column=match_col_idx).value
            
            if transaction_desc in category_map:
                category = category_map[transaction_desc].get(cat_col, '')
                subcategory = category_map[transaction_desc].get(subcat_col, '')
                
                # Update the cells in columns F and G
                sheet.cell(row=row_idx, column=cat_col_idx).value = category
                sheet.cell(row=row_idx, column=subcat_col_idx).value = subcategory

        # Step 6: Save the updated workbook.
        print(f"Saving changes directly to: '{master_file}'...")
        workbook.save(master_file)

        print("---")
        print("Script finished successfully! The original file has been updated.")
        print("Please check your file: 'master_transactions.xlsx'.")

    except FileNotFoundError as e:
        print(f"Error: File not found. Please check that '{e.filename}' exists in the current directory.", file=sys.stderr)
        sys.exit(1)
    except Exception as e:
        print(f"An unexpected error occurred: {e}", file=sys.stderr)
        sys.exit(1)

# Run the function
if __name__ == "__main__":
    categorize_transactions_with_openpyxl(MASTER_FILE, CATEGORIES_FILE, CATEGORIES_SHEET_NAME,
                                         MATCH_COLUMN, CATEGORY_COLUMN, SUBCATEGORY_COLUMN)

import pandas as pd
import os
import openpyxl
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font
from typing import Dict, List, Any
import re
from dateutil.relativedelta import relativedelta
from config import file_formats

# This script consolidates bank and credit card transactions from different CSV and XLSX formats
# into a single, standardized master file.

def process_transactions(file_path: str, format_info: Dict[str, Any]) -> tuple[pd.DataFrame, Dict[str, Any]]:
    """
    Reads a transaction file (CSV or XLSX), processes it, and returns a standardized DataFrame
    along with processing statistics.

    Args:
        file_path (str): The path to the input transaction file.
        format_info (Dict[str, Any]): The configuration dictionary for this file format.

    Returns:
        tuple[pd.DataFrame, Dict[str, Any]]: A tuple containing the standardized transaction data
        and a dictionary of processing statistics. Returns an empty DataFrame and empty dict on error.
    """
    corrected_month_date = None
    account_number = None
    installment_note_prefix = None
    
    processing_stats = {
        'file_name': os.path.basename(file_path),
        'original_rows': 0,
        'currency': format_info.get('currency', 'N/A'),
        'account_type': format_info.get('account_type', 'N/A'),
        'installment_rows': 0
    }

    try:
        header_row = format_info.get('header_row', 1)
        
        # Determine the file type and read it accordingly
        if file_path.endswith('.xlsx'):
            # For the Mastercard format, we need to read the first few rows separately
            if file_path.startswith('פירוט חיובים לכרטיס מאסטרקארד'):
                # Read cell A1 to extract the account number
                df_header = pd.read_excel(file_path, header=None, nrows=1)
                account_str = str(df_header.iloc[0, 0])
                account_match = re.search(r'(\d+)\s*$', account_str)
                if account_match:
                    account_number = int(account_match.group(1))

                # Read row 3 to get the billing date and note prefix for installments
                df_temp = pd.read_excel(file_path, header=None, nrows=3)
                billing_date_str = df_temp.iloc[2, 0]
                
                # Extract the date for the date logic
                date_match = re.search(r'ב-(\d{2}/\d{2}/\d{4})', billing_date_str)
                if date_match:
                    billing_date = pd.to_datetime(date_match.group(1), format='%d/%m/%Y')
                    
                    # Explicitly handle the month and year adjustment
                    target_month = billing_date.month - 1
                    target_year = billing_date.year
                    if target_month == 0:
                        target_month = 12
                        target_year -= 1
                    
                    corrected_month_date = billing_date.replace(year=target_year, month=target_month)
                
                # Extract the prefix for the Note column
                prefix_match = re.search(r'^(.*?)\:', billing_date_str)
                if prefix_match:
                    installment_note_prefix = prefix_match.group(1).strip()
                
                df = pd.read_excel(file_path, header=header_row - 1)
                
            else:
                df = pd.read_excel(file_path, header=header_row - 1)
                account_number = format_info.get('account_number', None)
        elif file_path.endswith('.csv'):
            df = pd.read_csv(file_path, header=header_row - 1)
            account_number = format_info.get('account_number', None)
        else:
            print(f"Error: Unsupported file type for '{file_path}'. Skipping.")
            return pd.DataFrame(), {}
            
    except FileNotFoundError:
        print(f"Error: The file '{file_path}' was not found.")
        return pd.DataFrame(), {}
    except Exception as e:
        print(f"Error reading file '{file_path}': {e}")
        return pd.DataFrame(), {}
    
    # Store the original number of rows before any filtering
    processing_stats['original_rows'] = len(df)
    
    # Clean up column names by stripping whitespace and replacing non-standard spaces
    df.columns = df.columns.str.strip()
    df.columns = df.columns.str.replace(r'\s+', ' ', regex=True)

    format_map = format_info['format_map']
    account_type = format_info['account_type']
    currency = format_info['currency']
    
    # Create a new, cleaned format map to ensure an exact match with the dataframe columns
    cleaned_format_map = {key.strip(): value for key, value in format_map.items()}

    # Rename columns based on the cleaned format map, using a more robust search
    # to find matching columns even with minor differences.
    df_columns = df.columns.tolist()
    new_column_names = {}
    for source_col, target_col in cleaned_format_map.items():
        # Find the actual column in the DataFrame that contains the source column name
        found_col = next((col for col in df_columns if source_col in col), None)
        if found_col:
            new_column_names[found_col] = target_col
    
    df = df.rename(columns=new_column_names)

    # Find the marker row before any type conversions
    stop_marker = 'תנועות עתידיות'
    try:
        if 'transaction_date' in df.columns:
            # Check for the stop marker in the 'transaction_date' column
            stop_indices = df[df['transaction_date'].str.contains(stop_marker, na=False, case=False)].index.tolist()
            if stop_indices:
                stop_index = stop_indices[0]
                df = df.loc[:stop_index - 1].copy()
    except KeyError:
        pass

    # Convert the transaction_date column to datetime objects
    df['transaction_date'] = pd.to_datetime(df['transaction_date'], dayfirst=True, errors='coerce')
    last_valid_date_index = df['transaction_date'].last_valid_index()
    if last_valid_date_index is not None and last_valid_date_index < len(df) - 1:
        df = df.iloc[:last_valid_date_index + 1]
    
    # Check for both 'debit_amount' and 'credit_amount' columns. If they exist, combine them.
    if 'debit_amount' in df.columns and 'credit_amount' in df.columns:
        # Fill NaN values with 0 and convert to numeric to enable calculations
        df['debit_amount'] = pd.to_numeric(df['debit_amount'].astype(str).str.replace(',', ''), errors='coerce').fillna(0)
        df['credit_amount'] = pd.to_numeric(df['credit_amount'].astype(str).str.replace(',', ''), errors='coerce').fillna(0)
        
        # Calculate original_amount: logic depends on account type
        if account_type == 'Credit Card':
            # For credit cards, a debit is a negative transaction
            df['original_amount'] = df['debit_amount'] - df['credit_amount']
        else:
            # For other accounts (current, etc.), a debit is a negative transaction
            df['original_amount'] = df['credit_amount'] - df['debit_amount']
    
    # Handle the case where there is a single 'original_amount' column
    elif 'original_amount' in df.columns:
        df['original_amount'] = pd.to_numeric(df['original_amount'].astype(str).str.replace(',', ''), errors='coerce').fillna(0)
        # If it's a credit card, make the amounts negative
        if account_type == 'Credit Card':
            df['original_amount'] = -df['original_amount']
    
    # Ensure all required columns exist in the DataFrame
    required_cols = list(set(format_map.values()))
    if not all(col in df.columns for col in required_cols if col not in ['debit_amount', 'credit_amount']):
        missing_cols = [col for col in required_cols if col not in df.columns and col not in ['debit_amount', 'credit_amount']]
        print(f"Error: Missing required columns in '{file_path}': {missing_cols}")
        return pd.DataFrame(), {}
    
    # Apply the installment date and note logic for Mastercard files
    if file_path.startswith('פירוט חיובים לכרטיס מאסטרקארד') and 'transaction_type' in df.columns:
        is_installment = df['transaction_type'].isin(['תשלומים', 'רכישה בקרדיט'])
        
        # Count the number of installment rows for the statistics
        processing_stats['installment_rows'] = is_installment.sum()

        if corrected_month_date:
            # Replace the month and year of 'Transaction Date' for installment rows
            df.loc[is_installment, 'transaction_date'] = df.loc[is_installment, 'transaction_date'].apply(
                lambda x: x.replace(month=corrected_month_date.month, year=corrected_month_date.year)
            )

        if installment_note_prefix:
            # Append the note prefix for installment rows
            df.loc[is_installment, 'Note'] = df.loc[is_installment, 'Note'].fillna('').astype(str) + ' ' + installment_note_prefix

    df['Transaction Account'] = account_number
    
    processing_stats['account_number'] = account_number
    
    df['Credit/Debit'] = df['original_amount']
    df['Category'] = ''
    df['Subcategory'] = ''
    
    # Check if 'Note' column exists after renaming. If not, add it as an empty column.
    if 'Note' not in df.columns:
        df['Note'] = ''
    
    df['Currency'] = currency

    # Keep only the required columns and reorder them.
    master_columns = [
        'transaction_date',
        'Transaction Account',
        'transaction_description',
        'Currency',
        'Credit/Debit',
        'Category',
        'Subcategory',
        'Note',
    ]
    
    final_df = df[master_columns]

    # Rename the date and description columns to their final names for the output file
    final_df = final_df.rename(columns={
        'transaction_date': 'Transaction Date',
        'transaction_description': 'Transaction Description'
    })

    return final_df, processing_stats


def consolidate_data(input_file_paths: List[str], output_file: str):
    """
    Consolidates transactions from a list of files and saves them to a master Excel file.

    Args:
        input_file_paths (List[str]): A list of file paths to process.
        output_file (str): The path to the output master Excel file.
    """
    all_dfs = []
    all_stats = []
    
    # Process each file in the list
    for file_path in input_file_paths:
        file_name = os.path.basename(file_path)
        
        # Check if the filename starts with any of the configured prefixes
        format_info = None
        for prefix in file_formats:
            if file_name.startswith(prefix):
                format_info = file_formats[prefix]
                
                # Check for a fixed account number in the format map. If not, use the prefix as account number
                if 'account_number' not in format_info:
                    try:
                        format_info['account_number'] = int(prefix)
                    except ValueError:
                        format_info['account_number'] = prefix

                print(f"Processing '{file_path}' using format '{prefix}'...")
                df, stats = process_transactions(file_path, format_info)
                if not df.empty:
                    all_dfs.append(df)
                    all_stats.append(stats)
                break
        
        if not format_info:
            print(f"Error: No configuration found for file '{file_name}'. Skipping.")

    if all_dfs:
        final_df = pd.DataFrame()
        seen_keys = {}
        all_original_rows = 0
        all_installment_rows = 0
        duplicate_report = {}
        
        # Process each DataFrame for deduplication and stats
        for df, stats in zip(all_dfs, all_stats):
            all_original_rows += stats['original_rows']
            all_installment_rows += stats['installment_rows']
            
            # Create a unique key for each transaction within the current file
            df['temp_key'] = df.apply(lambda row: (
                row['Transaction Date'].strftime('%Y-%m-%d') if pd.notna(row['Transaction Date']) else 'NaT',
                row['Transaction Account'],
                row['Transaction Description'],
                row['Credit/Debit']
            ), axis=1)

            # Find and process unique rows for the current file
            unique_rows_in_file = df.drop_duplicates(subset=['temp_key'])
            
            for index, row in unique_rows_in_file.iterrows():
                key = row['temp_key']
                
                if key in seen_keys:
                    # Found a duplicate from a previous file, record it
                    if key not in duplicate_report:
                        duplicate_report[key] = [seen_keys[key], stats['file_name']]
                    elif stats['file_name'] not in duplicate_report[key]:
                        duplicate_report[key].append(stats['file_name'])
                else:
                    # This is a new, unique transaction across all files processed so far
                    seen_keys[key] = stats['file_name']
                    final_df = pd.concat([final_df, pd.DataFrame([row])], ignore_index=True)

        final_df = final_df.drop(columns='temp_key')
        final_count = len(final_df)

        try:
            # Use pandas to save the dataframe to an Excel file using the openpyxl engine
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                final_df.to_excel(writer, index=False, sheet_name='Transactions')

            # Now, we use openpyxl to apply formatting since pandas doesn't support it directly
            workbook = openpyxl.load_workbook(output_file)
            worksheet = workbook.active
            
            # Get the column letter for 'Transaction Date'
            date_col_letter = None
            for cell in worksheet[1]: # Iterate through the header row
                if cell.value == 'Transaction Date':
                    date_col_letter = cell.column_letter
                    break
            
            # Format the 'Transaction Date' column as a Short Date
            if date_col_letter:
                for row_num in range(2, worksheet.max_row + 1):
                    cell = worksheet[f'{date_col_letter}{row_num}']
                    cell.number_format = 'dd/mm/yyyy'

            # Get the column letter for 'Credit/Debit'
            credit_debit_col_letter = None
            for cell in worksheet[1]: # Iterate through the header row
                if cell.value == 'Credit/Debit':
                    credit_debit_col_letter = cell.column_letter
                    break

            if credit_debit_col_letter:
                # Define the font colors
                red_font = Font(color="FFFF0000")
                green_font = Font(color="FF00B050")
                
                # Apply conditional formatting for negative numbers (red font)
                worksheet.conditional_formatting.add(
                    f'{credit_debit_col_letter}2:{credit_debit_col_letter}{worksheet.max_row}',
                    CellIsRule(operator='lessThan', formula=['0'], font=red_font)
                )

                # Apply conditional formatting for positive numbers (green font)
                worksheet.conditional_formatting.add(
                    f'{credit_debit_col_letter}2:{credit_debit_col_letter}{worksheet.max_row}',
                    CellIsRule(operator='greaterThan', formula=['0'], font=green_font)
                )

            
            # Auto-fit all columns with a more robust calculation
            for column in worksheet.columns:
                max_length = 0
                column_letter = openpyxl.utils.get_column_letter(column[0].column)
                for cell in column:
                    try:
                        cell_content = str(cell.value) if cell.value is not None else ""
                        if len(cell_content) > max_length:
                            max_length = len(cell_content)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column_letter].width = adjusted_width

            workbook.save(output_file)
            
            print(f"\nSuccessfully consolidated and formatted transactions to '{output_file}'!")
        
        # New error handling for when the file is open
        except PermissionError:
            print(f"Error: The output file '{output_file}' is currently open in another program. Please close the file and try again.")
            
        except Exception as e:
            print(f"Error: An unexpected error occurred while saving the Excel file: {e}")
            print("Falling back to saving as CSV.")
            final_df.to_csv('master_transactions.csv', index=False)
            print(f"Successfully consolidated transactions to 'master_transactions.csv' instead.")

        # Print all the new summary information
        print("\n--- File Processing Summary ---")
        for stats in all_stats:
            print(f"\nFile: {stats['file_name']}")
            print(f"  Account Number: {stats.get('account_number', 'N/A')}")
            print(f"  Account Type: {stats['account_type']}")
            print(f"  Currency: {stats['currency']}")
            print(f"  Original Rows: {stats['original_rows']}")
            print(f"  Installment Rows: {stats['installment_rows']}")
        
        print("\n--- Overall Summary ---")
        print(f"  Total Original Rows: {all_original_rows}")
        print(f"  Total Consolidated Rows: {final_count}")
        print(f"  Total Duplicated Rows: {len(duplicate_report)}")
        print(f"  Total Installment Rows: {all_installment_rows}")

        if duplicate_report:
            print("\n--- Duplicate Transaction Report ---")
            for key, files in duplicate_report.items():
                key_tuple_str = str(key).replace("'", "")
                print(f"  Duplicate found: {key_tuple_str}")
                print(f"    - Found in files: {', '.join(files)}")

    else:
        print("\nNo transactions were processed. The output file was not created.")


if __name__ == '__main__':
    # --- Example Usage ---
    # Find all CSV or XLSX files in the current directory that have a configured prefix
    input_files = []
    for filename in os.listdir('.'):
        if filename.endswith(('.csv', '.xlsx')):
            input_files.append(filename)

    if not input_files:
        print("No CSV or XLSX files found with a configured prefix. Please ensure your files are in the same directory.")
    else:
        consolidate_data(input_files, 'master_transactions.xlsx')
